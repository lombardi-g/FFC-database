from openpyxl import load_workbook

excel_file = "Banco de Dados Figueirense Base.xlsx"

def pass_to_excel(date,category,tournament,figueira_final_score,opponent_final_score,opponent,home,place,city,first_half_minutes,second_half_minutes,figueira_first,scored_list,conceded_list,minutes_list):

    workbook = load_workbook(excel_file)
    sheet = workbook['Jogos']
    # TODO:Hard-coded labels from sheet. Read them and dynamically make the dict?
    column_labels = {
        "CÓDIGO JOGO":1,
        "DATA JOGO":2,
        "TREINADOR":3,
        "CATEGORIA":4,
        "COMPETIÇÃO":5,
        "FIGUEIRENSE":6,
        "G.F.":7,
        "G.A.":8,
        "ADVERSÁRIO":9,
        "MANDO":10,
        "LOCAL":11,
        "Cidade":12,
        "UF":13,
        "JOGOS":14,
        "VITÓRIA":15,
        "EMPATE":16,
        "DERROTA":17,
        "MINUTOS JOGADOS":18,
        "1º A MARCAR FIGUEIRENSE":19,
        "1º A MARCAR ADVERSÁRIO":20,
        "GOLS MARCADOS 1ºT - 0'-15'":21,
        "GOLS MARCADOS 1ºT - 15'-30'":22,
        "GOLS MARCADOS 1ºT - 30'-45'":23,
        "GOLS MARCADOS 2ºT - 0'-15'":24,
        "GOLS MARCADOS 2ºT - 15'-30'":25,
        "GOLS MARCADOS 2ºT - 30'-45'":26,
        "GOLS SOFRIDOS 1ºT - 0'-15'":27,
        "GOLS SOFRIDOS 1ºT - 15'-30'":28,
        "GOLS SOFRIDOS 1ºT - 30'-45'":29,
        "GOLS SOFRIDOS 2ºT - 0'-15'":30,
        "GOLS SOFRIDOS 2ºT - 15'-30'":31,
        "GOLS SOFRIDOS 2ºT - 30'-45'":32
    }
    staff = {
        "Sub14":"Guilherme Pereira",
        "Sub15":"Guilherme Pereira",
        "Sub17":"Lucas Ligio",
        "Sub20":"Jhonatas Reis",
        "Sub21":"Jhonatas Reis"
    }

    last_row = sheet.max_row
    last_row_value = 2
    while last_row > 1:
        last_row_value = sheet.cell(row=last_row,column=1).value
        if last_row_value is not None:
            break
        last_row -= 1
        
    last_row_value = sheet.cell(row=last_row, column=column_labels['CÓDIGO JOGO']).value
    
    new_row = last_row + 1
    sheet.cell(row=new_row, column=column_labels['CÓDIGO JOGO'],value=last_row_value+1)
    sheet.cell(row=new_row, column=column_labels['DATA JOGO'],value=date)
    sheet.cell(row=new_row, column=column_labels['TREINADOR'],value=staff[category])
    sheet.cell(row=new_row, column=column_labels['CATEGORIA'],value=category)
    sheet.cell(row=new_row, column=column_labels['COMPETIÇÃO'],value=tournament)
    sheet.cell(row=new_row, column=column_labels['FIGUEIRENSE'],value="Figueirense")
    sheet.cell(row=new_row, column=column_labels['G.F.'],value=figueira_final_score)
    sheet.cell(row=new_row, column=column_labels['G.A.'],value=opponent_final_score)
    sheet.cell(row=new_row, column=column_labels['ADVERSÁRIO'],value=opponent)
    sheet.cell(row=new_row, column=column_labels['MANDO'],value=home)
    sheet.cell(row=new_row, column=column_labels['LOCAL'],value=place)
    sheet.cell(row=new_row, column=column_labels['Cidade'],value=city)
    sheet.cell(row=new_row, column=column_labels['UF'],value="SC")
    sheet.cell(row=new_row, column=column_labels['JOGOS'],value=1)
    sheet.cell(row=new_row, column=column_labels['VITÓRIA'],value=1 if figueira_final_score > opponent_final_score else 0)
    sheet.cell(row=new_row, column=column_labels['EMPATE'],value=1 if figueira_final_score == opponent_final_score else 0)
    sheet.cell(row=new_row, column=column_labels['DERROTA'],value=1 if figueira_final_score < opponent_final_score else 0)
    sheet.cell(row=new_row, column=column_labels['MINUTOS JOGADOS'],value = int(first_half_minutes) + int(second_half_minutes))
    sheet.cell(row=new_row, column=column_labels['1º A MARCAR FIGUEIRENSE'],value=1 if figueira_first == True else 0)
    sheet.cell(row=new_row, column=column_labels['1º A MARCAR ADVERSÁRIO'],value=1 if figueira_first == False else 0)
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 1ºT - 0\'-15\''], value= scored_list[0])
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 1ºT - 15\'-30\''],value= scored_list[1])
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 1ºT - 30\'-45\''],value= scored_list[2])
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 2ºT - 0\'-15\''], value= scored_list[3])
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 2ºT - 15\'-30\''],value= scored_list[4])
    sheet.cell(row=new_row, column=column_labels['GOLS MARCADOS 2ºT - 30\'-45\''],value= scored_list[5])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 1ºT - 0\'-15\''], value= conceded_list[0])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 1ºT - 15\'-30\''],value= conceded_list[1])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 1ºT - 30\'-45\''],value= conceded_list[2])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 2ºT - 0\'-15\''], value= conceded_list[3])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 2ºT - 15\'-30\''],value= conceded_list[4])
    sheet.cell(row=new_row, column=column_labels['GOLS SOFRIDOS 2ºT - 30\'-45\''],value= conceded_list[5])

    # Minute scraping
    sheet = workbook['Scout']

    column_labels = {
        "Data":1,
        "Competição": 3,
        "Adversário":4,
        "Nome":8,
        "Minutos":17        
    }

    last_row = sheet.max_row
    last_row_value = 2
    while last_row > 1:
        last_row_value = sheet.cell(row=last_row,column=1).value
        if last_row_value is not None:
            break
        last_row -= 1
    new_row = last_row + 1

    for player in minutes_list:
        sheet.cell(row=new_row, column=column_labels['Data'],value = date)
        sheet.cell(row=new_row, column=column_labels['Competição'],value = category)
        sheet.cell(row=new_row, column=column_labels['Adversário'],value = opponent)
        sheet.cell(row=new_row, column=column_labels['Nome'],value = player["name"])
        sheet.cell(row=new_row, column=column_labels['Minutos'],value = player["minutes_played"])
        new_row += 1

    workbook.save(excel_file)